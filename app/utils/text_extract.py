# app/utils/text_extract.py
import re
from pathlib import Path
import mimetypes
import pymupdf as fitz          # PDF
import docx
import email
from email import policy
import extract_msg              # Outlook .msg
import openpyxl                 # Excel
from app.core.logger import logger


# ============================================================================
# MAIN FILE ROUTER
# ============================================================================
def extract_text(file_path: str) -> str:
    """
    Extrae texto desde:
      ✔ PDF
      ✔ DOCX
      ✔ TXT
      ✔ XLSX
      ✔ EML
      ✔ MSG (Outlook)
      ✔ Imágenes → OCR desactivado (listo para OpenAI Vision)
    """
    path = Path(file_path)
    suffix = path.suffix.lower()
    mime, _ = mimetypes.guess_type(path)

    try:
        # -------------------------------------------------------------
        # PDF
        # -------------------------------------------------------------
        if mime == "application/pdf" or suffix == ".pdf":
            logger.info(f"📄 Extrayendo PDF: {path}")
            return extract_text_pdf(path)

        # -------------------------------------------------------------
        # DOCX
        # -------------------------------------------------------------
        if suffix == ".docx":
            logger.info(f"📝 Extrayendo DOCX: {path}")
            return extract_text_docx(path)

        # -------------------------------------------------------------
        # TXT
        # -------------------------------------------------------------
        if mime == "text/plain" or suffix == ".txt":
            logger.info(f"📄 Extrayendo TXT: {path}")
            return extract_text_txt(path)

        # -------------------------------------------------------------
        # EXCEL
        # -------------------------------------------------------------
        if suffix == ".xlsx":
            logger.info(f"📊 Extrayendo Excel XLSX: {path}")
            return extract_text_excel(path)

        # -------------------------------------------------------------
        # EMAIL .EML
        # -------------------------------------------------------------
        if suffix == ".eml":
            logger.info(f"📧 Extrayendo correo EML: {path}")
            return extract_text_eml(path)

        # -------------------------------------------------------------
        # EMAIL .MSG (Outlook)
        # -------------------------------------------------------------
        if suffix == ".msg":
            logger.info(f"📨 Extrayendo correo MSG (Outlook): {path}")
            return extract_text_msg(path)

        # -------------------------------------------------------------
        # IMÁGENES (OCR OFF)
        # -------------------------------------------------------------
        if suffix in [".png", ".jpg", ".jpeg", ".tiff", ".bmp"]:
            logger.info(f"🖼 Imagen detectada, OCR deshabilitado: {path}")
            return "[Imagen detectada: OCR deshabilitado → OpenAI Vision disponible]"

        # -------------------------------------------------------------
        logger.warning(f"⚠ Tipo de archivo NO soportado: {suffix}")
        return ""

    except Exception as e:
        logger.error(f"❌ Error extrayendo texto de {file_path}: {e}")
        return ""


# ============================================================================
# PDF — PyMuPDF
# ============================================================================
def extract_text_pdf(path: Path) -> str:
    doc = fitz.open(path)
    final_text = []

    try:
        for page in doc:
            # Primera pasada — texto normal
            text1 = page.get_text("text")

            # Segunda pasada — bloques (más robusto)
            blocks = page.get_text("blocks")
            text2 = "\n".join([b[4] for b in blocks if isinstance(b, tuple) and len(b) > 4])

            # Escoger el más largo (mejor extracción)
            best = text1 if len(text1) > len(text2) else text2

            final_text.append(best)

    finally:
        doc.close()

    return clean_text("\n".join(final_text))


# ============================================================================
# DOCX
# ============================================================================
def extract_text_docx(path: Path) -> str:
    doc = docx.Document(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return clean_text("\n".join(paragraphs))


# ============================================================================
# TXT
# ============================================================================
def extract_text_txt(path: Path) -> str:
    return clean_text(path.read_text(encoding="utf-8", errors="ignore"))


# ============================================================================
# EXCEL (.xlsx)
# ============================================================================
def extract_text_excel(path: Path) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    content = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        content.append(f"\n--- HOJA: {sheet} ---\n")

        for row in ws.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None]
            if values:
                content.append(" | ".join(values))

    return clean_text("\n".join(content))


# ============================================================================
# EMAILS — formato .eml
# ============================================================================
def extract_text_eml(path: Path) -> str:
    with open(path, "rb") as f:
        msg = email.message_from_binary_file(f, policy=policy.default)

    subject = msg.get("subject", "")
    body = []

    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                try:
                    body.append(
                        part.get_payload(decode=True).decode("utf-8", errors="ignore")
                    )
                except Exception:
                    pass
    else:
        body.append(msg.get_payload(decode=True).decode("utf-8", errors="ignore"))

    text = f"""
ASUNTO: {subject}

CUERPO:
{''.join(body)}
"""
    return clean_text(text)


# ============================================================================
# EMAILS — formato .msg (Outlook)
# ============================================================================
def extract_text_msg(path: Path) -> str:
    msg = extract_msg.Message(str(path))

    text = f"""
DE: {msg.sender}
PARA: {msg.to}
ASUNTO: {msg.subject}

{msg.body}
"""
    return clean_text(text)


# ============================================================================
# TEXT CLEANER
# ============================================================================
def clean_text(text: str) -> str:
    # Mantengo saltos de línea porque son útiles para el chunker.
    text = text.replace("\r", "\n")
    # Normalizar múltiple espacios
    text = re.sub(r"[ \t]{2,}", " ", text)
    # Convertir múltiples saltos de línea en uno solo
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()
